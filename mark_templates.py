"""
템플릿 파일에 DB 출처를 표시한 마크업 버전을 생성합니다.
- commence_body.hwpx → commence_body_MARKED.hwpx
- design_estimate_template.xlsx → design_estimate_template_MARKED.xlsx
"""
import os
import shutil
import zipfile
import re

ROOT = r'C:\Users\PUJ\eclipse-workspace\swmanager\src\main\resources'
OUT_DIR = r'C:\Users\PUJ\eclipse-workspace\swmanager\template_review'
os.makedirs(OUT_DIR, exist_ok=True)

# ============================================================
# 1. HWPX 착수계 템플릿 마크업
# ============================================================
HWPX_SRC = os.path.join(ROOT, 'templates', 'hwpx', 'commence_body.hwpx')
HWPX_DST = os.path.join(OUT_DIR, '착수계_commence_body_MARKED.hwpx')

# 플레이스홀더 → DB 출처 매핑
HWPX_DB_MAP = {
    # 공문/본문 (sw_pjt + 사용자 입력)
    '용역명': 'sw_pjt.proj_nm',
    '수신': 'sw_pjt.city_nm+dist_nm+org_nm',
    '계약금액': 'sw_pjt.cont_amt',
    '계약기간': 'sw_pjt.cont_dt~end_dt',
    '제출일자': '[입력]letter.date',
    '용역기간': 'sw_pjt.cont_dt~end_dt',
    '제품명': 'sw_pjt.proj_nm',
    '작성년월': '[입력]',
    '착수일자_한글': 'sw_pjt.start_dt',
    # 공정명 - 마스터테이블
    '공정명': '★tb_process_master.process_name',
    '점검방법': '★tb_service_purpose.purpose_text(SCOPE)',
    # 현장대리인 (users + tb_contract_participant.is_site_rep=true)
    '현대_주소': 'users.addr',
    '현대_성명': 'users.username',
    '현대_주민번호': 'users.ssn',
    '현대_생년': 'users.ssn[0:6]',
    # 보안각서 (참여자 1~3)
    '보안1_성명': 'users.username',
    '보안2_성명': 'users.username',
    '보안3_성명': 'users.username',
    '보안1_등급': '★users.tech_grade',
    '보안2_등급': '★users.tech_grade',
    '보안3_등급': '★users.tech_grade',
    '보안1_업무': '★tb_contract_participant.task_desc',
    '보안2_업무': '★tb_contract_participant.task_desc',
    '보안3_업무': '★tb_contract_participant.task_desc',
    '보안1_주민번호': 'users.ssn',
    '보안2_주민번호': 'users.ssn',
    '보안3_주민번호': 'users.ssn',
    # 과업참여자
    '참여1_직위': 'users.position_title',
    '참여2_직위': 'users.position_title',
    '참여3_직위': 'users.position_title',
    '참여1_자격': 'users.certificate',
    '참여2_자격': 'users.certificate',
    '참여3_자격': 'users.certificate',
    # 보안서약서
    '서약1_직위': 'users.position_title',
    '서약2_직위': 'users.position_title',
    '서약3_직위': 'users.position_title',
    '서약1_성명': 'users.username',
    '서약2_성명': 'users.username',
    '서약3_성명': 'users.username',
    # 투입인력
    '인력1_성명': 'users.username',
    '인력2_성명': 'users.username',
    '인력3_성명': 'users.username',
    '인력1_직급': 'users.position_title',
    '인력2_직급': 'users.position_title',
    '인력3_직급': 'users.position_title',
    '인력1_등급': '★users.tech_grade',
    '인력2_등급': '★users.tech_grade',
    '인력3_등급': '★users.tech_grade',
}


def mark_hwpx():
    if os.path.exists(HWPX_DST):
        os.remove(HWPX_DST)

    with zipfile.ZipFile(HWPX_SRC, 'r') as src_zip:
        with zipfile.ZipFile(HWPX_DST, 'w', zipfile.ZIP_DEFLATED) as dst_zip:
            for entry in src_zip.namelist():
                data = src_zip.read(entry)

                if entry.endswith('.xml') or entry.endswith('.txt'):
                    try:
                        text = data.decode('utf-8')
                        # 각 플레이스홀더를 [DB:출처] 표기로 변환
                        for ph, src in HWPX_DB_MAP.items():
                            old = '{{' + ph + '}}'
                            new = '【' + ph + '◀' + src + '】'
                            text = text.replace(old, new)
                        data = text.encode('utf-8')
                    except UnicodeDecodeError:
                        pass

                # mimetype은 STORED 방식
                if entry == 'mimetype':
                    info = zipfile.ZipInfo(entry)
                    info.compress_type = zipfile.ZIP_STORED
                    dst_zip.writestr(info, data)
                else:
                    dst_zip.writestr(entry, data)

    print('[OK] HWPX 마크업 완료:', HWPX_DST)


# ============================================================
# 2. Excel 설계내역서 템플릿 마크업
# ============================================================
XLSX_SRC = os.path.join(ROOT, 'templates', 'excel', 'design_estimate_template.xlsx')
XLSX_DST = os.path.join(OUT_DIR, '설계내역서_design_estimate_template_MARKED.xlsx')


def mark_xlsx():
    import openpyxl
    from openpyxl.styles import PatternFill, Font
    from openpyxl.comments import Comment

    shutil.copy(XLSX_SRC, XLSX_DST)
    wb = openpyxl.load_workbook(XLSX_DST)

    yellow = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    orange = PatternFill(start_color='FFCC99', end_color='FFCC99', fill_type='solid')
    cyan = PatternFill(start_color='CCFFFF', end_color='CCFFFF', fill_type='solid')

    # ===== 표지 시트 =====
    if len(wb.sheetnames) >= 1:
        ws = wb.worksheets[0]
        marks_cover = [
            ('A4', '★DB: sw_pjt.year', yellow),
            ('A5', '★DB: sw_pjt.proj_nm', yellow),
            ('A15', '[입력] 설계일 (designDate)', cyan),
            ('F20', '★DB: sw_pjt.dist_nm', yellow),
        ]
        for cell_ref, comment_text, fill in marks_cover:
            try:
                c = ws[cell_ref]
                c.fill = fill
                c.comment = Comment(comment_text, 'DB-MAP')
            except Exception as e:
                print(f'  표지 {cell_ref} skip: {e}')

    # ===== 갑지 시트 =====
    if len(wb.sheetnames) >= 2:
        ws = wb.worksheets[1]
        marks_gapji = [
            ('B1', '[입력] 설계년월일 (designDate)', cyan),
            ('A5', '수식 참조: 표지!A5 (sw_pjt.proj_nm)', orange),
            ('C6', '[입력] 위치 (location)', cyan),
            ('C8', '[수식] 용역개요 자동', orange),
            ('C13', '[수식] 총용역비 한글금액', orange),
        ]
        for cell_ref, comment_text, fill in marks_gapji:
            try:
                c = ws[cell_ref]
                c.fill = fill
                c.comment = Comment(comment_text, 'DB-MAP')
            except Exception as e:
                print(f'  갑지 {cell_ref} skip: {e}')

    # ===== 총괄표 시트 =====
    if len(wb.sheetnames) >= 3:
        ws = wb.worksheets[2]
        marks_summary = [
            # HW 항목 (Excel row 5,6 = openpyxl row 5,6)
            ('B5', '★[입력]items[0].name (HW1 품명)', yellow),
            ('D5', '★[입력]items[0].rate (HW1 적용율)', yellow),
            ('E5', '★[입력]items[0].unitPrice (HW1 도입가)', yellow),
            ('B6', '★[입력]items[1].name (HW2 품명)', yellow),
            ('D6', '★[입력]items[1].rate (HW2 적용율)', yellow),
            ('E6', '★[입력]items[1].unitPrice (HW2 도입가)', yellow),
            # SW 항목
            ('B9', '★[입력]items[].name (SW1 품명)', yellow),
            ('D9', '★[입력]items[].rate (SW1 적용율)', yellow),
            ('E9', '★[입력]items[].unitPrice (SW1 도입가)', yellow),
            ('B10', '★[입력]items[].name (SW2 품명)', yellow),
            ('D10', '★[입력]items[].rate (SW2 적용율)', yellow),
            ('E10', '★[입력]items[].unitPrice (SW2 도입가)', yellow),
            # 수식 셀
            ('C5', '[수식] 적용율 텍스트', orange),
            ('F5', '[수식] 도입가*적용율', orange),
            ('G5', '[수식] 경비', orange),
            ('H5', '[수식] 유지관리 소계', orange),
            ('H11', '[수식] ROUNDDOWN 낙찰율(0.9696)', orange),
        ]
        for cell_ref, comment_text, fill in marks_summary:
            try:
                c = ws[cell_ref]
                c.fill = fill
                c.comment = Comment(comment_text, 'DB-MAP')
            except Exception as e:
                print(f'  총괄표 {cell_ref} skip: {e}')

    wb.save(XLSX_DST)
    print('[OK] Excel 마크업 완료:', XLSX_DST)


# ============================================================
# 3. README
# ============================================================
README = """\
# 템플릿 DB 매핑 검토용 파일

## 표시 규칙

### HWPX (착수계)
원본의 `{{플레이스홀더}}` 부분을 `【플레이스홀더◀DB출처】` 형식으로 치환했습니다.
한글에서 열어 각 항목의 출처를 확인하세요.

- `★` 표시: 별도 마스터 테이블 또는 추가 입력이 필요한 항목
- `[입력]` 표시: DB가 아닌 사용자가 폼에서 직접 입력하는 항목

### Excel (설계내역서)
셀 배경색과 셀 메모(comment)로 출처를 표시했습니다.
- 노란색 = DB 또는 사용자 입력 (덮어쓰기 대상)
- 시안색 = 사용자 입력
- 주황색 = 수식 (자동 계산)

각 셀에 마우스 올리면 출처 메모가 표시됩니다.

## 검토 후
원하는 셀/플레이스홀더에 어떤 DB 출처를 추가하고 싶은지 알려주세요.
"""

readme_path = os.path.join(OUT_DIR, 'README.md')
with open(readme_path, 'w', encoding='utf-8') as f:
    f.write(README)

if __name__ == '__main__':
    mark_hwpx()
    mark_xlsx()
    print('\n생성 위치:', OUT_DIR)
